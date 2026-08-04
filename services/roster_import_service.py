"""球员库 Excel/CSV 导入服务。

- xlsx 使用 openpyxl 流式读取（read_only）；csv 自动探测编码（utf-8-sig → gbk → utf-8）
- 按配置列位取值，数值单元格防精度变形
- 分批事务 upsert（按 player_uid 覆盖更新）
- imports 目录文件数超上限时自动删除最旧
"""

import csv
import os
from pathlib import Path

from astrbot.api import logger

from ..utils.security import sanitize_text, sanitize_uid

_ALLOWED_EXTS = (".xlsx", ".csv")


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return format(value, ".15f").rstrip("0").rstrip(".")
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _parse_csv(file_path: Path, max_rows: int):
    raw = file_path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = raw.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    rows = []
    for i, line in enumerate(csv.reader(text.splitlines())):
        if i > max_rows:
            break
        rows.append(line)
    return rows


def _parse_xlsx(file_path: Path, max_rows: int):
    from openpyxl import load_workbook

    rows = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > max_rows:
                break
            rows.append([_cell_to_str(v) for v in row])
    finally:
        wb.close()
    return rows


class RosterImportService:
    def __init__(self, db, dao, cfg):
        self._db = db
        self._dao = dao
        self._cfg = cfg

    @property
    def imports_dir(self) -> Path:
        base = Path(self._db.db_path).parent
        d = base / "imports"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _col(self, key: str) -> int:
        return int(self._cfg.get(key, 0))

    def list_files(self) -> list:
        files = [
            p for p in self.imports_dir.iterdir()
            if p.is_file() and p.suffix.lower() in _ALLOWED_EXTS
        ]
        files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return files

    def check_file(self, file_name: str) -> Path:
        from ..utils.security import sanitize_filename

        safe = sanitize_filename(file_name)
        p = self.imports_dir / safe
        if not p.is_file() or p.parent.resolve() != self.imports_dir.resolve():
            raise FileNotFoundError(f"文件「{file_name}」不存在于导入目录")
        if p.suffix.lower() not in _ALLOWED_EXTS:
            raise ValueError("仅支持 .xlsx / .csv 文件")
        size_mb = int(self._cfg.get("import_max_file_size_mb", 50))
        if p.stat().st_size > size_mb * 1024 * 1024:
            raise ValueError(f"文件超过大小上限（{size_mb} MB）")
        return p

    def cleanup_oldest(self) -> int:
        max_files = int(self._cfg.get("import_max_files", 50))
        files = self.list_files()
        removed = 0
        while len(files) > max_files:
            old = files.pop()
            try:
                old.unlink()
                removed += 1
            except OSError as e:
                logger.warning(f"Failed to remove import file {old}: {e}")
        return removed

    def parse_rows(self, file_path: Path) -> tuple[list, list, int]:
        """解析文件，返回 (数据行, 错误信息, 跳过的表头/空行数)。

        数据行格式: (player_uid, foreign_name, chinese_name)
        """
        max_rows = int(self._cfg.get("import_max_rows", 50000))
        if file_path.suffix.lower() == ".xlsx":
            raw_rows = _parse_xlsx(file_path, max_rows)
        else:
            raw_rows = _parse_csv(file_path, max_rows)

        col_uid = self._col("import_col_uid")
        col_foreign = self._col("import_col_foreign_name")
        col_cn = self._col("import_col_chinese_name")

        data: list[tuple[str, str, str]] = []
        errors: list[str] = []
        skipped = 0
        for idx, row in enumerate(raw_rows, start=1):
            if not row or not any(str(c).strip() for c in row):
                skipped += 1
                continue
            uid_raw = _cell_to_str(row[col_uid - 1]) if 0 < col_uid <= len(row) else ""
            foreign_raw = (
                _cell_to_str(row[col_foreign - 1]) if 0 < col_foreign <= len(row) else ""
            )
            cn_raw = _cell_to_str(row[col_cn - 1]) if 0 < col_cn <= len(row) else ""

            uid = sanitize_uid(uid_raw)
            foreign = sanitize_text(foreign_raw)
            cn = sanitize_text(cn_raw)
            if not uid:
                errors.append(f"第{idx}行: 球员 ID 为空或非法")
                continue
            if not foreign:
                errors.append(f"第{idx}行: 外文名为空")
                continue
            data.append((uid, foreign, cn))
        return data, errors, skipped

    async def import_rows(self, data: list[tuple[str, str, str]], created_by: str) -> dict:
        """分批事务导入，返回统计。"""
        batch_size = max(1, int(self._cfg.get("import_batch_size", 5000)))
        added = 0
        updated = 0

        def _chunks():
            for i in range(0, len(data), batch_size):
                yield data[i : i + batch_size]

        for chunk in _chunks():
            async def _tx(conn):
                add = 0
                upd = 0
                for uid, foreign, cn in chunk:
                    async with conn.execute(
                        "SELECT 1 FROM player_roster WHERE player_uid=?", (uid,)
                    ) as cur:
                        exists = await cur.fetchone()
                    await self._dao.upsert_player(conn, uid, foreign, cn, created_by)
                    if exists:
                        upd += 1
                    else:
                        add += 1
                return add, upd

            a, u = await self._db.execute_transaction(_tx)
            added += a
            updated += u
        return {"added": added, "updated": updated}

    def preview(self, file_path: Path, limit: int = 3) -> list:
        data, _, _ = self.parse_rows(file_path)
        return data[:limit]

    def save_uploaded(self, file_path: str, file_name: str) -> Path:
        """将已下载的群发文件移入导入目录（触发自动清理）。"""
        from ..utils.security import sanitize_filename

        safe_name = sanitize_filename(file_name)
        if Path(safe_name).suffix.lower() not in _ALLOWED_EXTS:
            raise ValueError("仅支持 .xlsx / .csv 文件")
        size_mb = int(self._cfg.get("import_max_file_size_mb", 50))
        size = os.path.getsize(file_path)
        if size > size_mb * 1024 * 1024:
            raise ValueError(f"文件超过大小上限（{size_mb} MB）")
        target = self.imports_dir / safe_name
        target.write_bytes(Path(file_path).read_bytes())
        self.cleanup_oldest()
        return target
